# app.py

from flask import Flask, request, jsonify
from flask_cors import CORS
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
import json
import os
import logging
from datetime import datetime
import pandas as pd
import glob
import time
import random
from bs4 import BeautifulSoup
import requests
import re
from functools import wraps
from pymongo import MongoClient
import uuid
from werkzeug.security import generate_password_hash, check_password_hash
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
import pandas as pd
import io
from flask import send_file
from openpyxl import load_workbook
from urllib.parse import quote
from urllib.parse import quote
import random
from openpyxl import load_workbook
import io
from flask import send_file
from urllib.parse import quote
import random
from collections import Counter
import ssl
import traceback


# Flask 앱 초기화
app = Flask(__name__)
CORS(app, resources={r"/*": {"origins": "*"}}, supports_credentials=True)



# MongoDB 연결 정보 수정
uri = "mongodb+srv://shinws8908:dnfhlao1@cluster0.h7c55.mongodb.net/?retryWrites=true&w=majority&appName=Cluster0"
client = MongoClient(uri)
db = client['scraping_db']

# 로깅 설정
logging.basicConfig(level=logging.DEBUG)
logger = logging.getLogger(__name__)

# 구성 설정
config = {
    'Paths': {
        'download_folder': r"C:\Downloads\네이버쇼핑데이터",
        'processed_data': r"C:\Downloads\네이버쇼핑데이터\processed_data"
    },
    'Settings': {
        'excluded_markets': ['excluded1', 'excluded2']
    }
}

# 디렉토리 생성
os.makedirs(config['Paths']['download_folder'], exist_ok=True)
os.makedirs(config['Paths']['processed_data'], exist_ok=True)

# JSON 데이터 로드 함수
def load_json_data(file_path):
    try:
        if os.path.exists(file_path) and os.path.getsize(file_path) > 0:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return data if isinstance(data, list) else []
        return []
    except json.JSONDecodeError:
        logger.error(f"JSON 디코딩 오류: {file_path}")
        return []


def save_json_data(file_path, data):
    with open(file_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=4)


# 네이버 쇼핑 스크래핑 클래스
class NaverShoppingScraper:
    # 초기화
    def __init__(self, user_folder):
        self.driver = None
        self.user_folder = user_folder
        self.config_path = os.path.join(user_folder, 'config.json')
        self.config = self.load_or_create_config()

    # 설정파일 로드 또는 생성
    def load_or_create_config(self):
        default_config = {
            "market": 0,
            "min_price": 0,
            "max_price": 1000000000,
            "option": "전체구매건수",
            "skip_words": [],
            "markets": []
        }
        if not os.path.exists(self.config_path):
            with open(self.config_path, 'w', encoding='utf-8') as f:
                json.dump(default_config, f, ensure_ascii=False, indent=4)
            logger.info(f"기본 설정 파일 생성됨: {self.config_path}")
            return default_config
        else:
            try:
                with open(self.config_path, 'r', encoding='utf-8') as f:
                    loaded_config = json.load(f)
                if not isinstance(loaded_config, dict):
                    logger.warning("로드된 설정이 딕셔너리가 닙니다. 기  사용합니다.")
                    return default_config
                logger.info(f"사용자 설정 로드 완료: {loaded_config}")
                return loaded_config
            except json.JSONDecodeError as e:
                logger.error(f"설정 파일 파싱 류: {str(e)}")
                return default_config

    # 웹드라이버 설정
    def setup_driver(self):
        chrome_options = Options()
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--disable-blink-features=AutomationControlled")
        chrome_options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/90.0.4430.212 Safari/537.36")
        chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
        chrome_options.add_experimental_option('useAutomationExtension', False)
        
        try:
            service = Service(ChromeDriverManager().install())
            self.driver = webdriver.Chrome(
                service=service,
                options=chrome_options
            )
            self.wait = WebDriverWait(self.driver, 30)
            logger.debug("WebDriver 설정 완료")
        except Exception as e:
            logger.error(f"WebDriver 설정 중 오류 발생: {str(e)}")
            self.driver = None
            raise

    # 상품 제목 필터링
    def should_skip_title(self, title):
        title_words = title.split()
        for word in title_words:
            if word in self.config.get('skip_words', []):
                return True
        if re.search(r"[^\w\s-]", title):
            return True
        return False

    # 스마트스토어 상품 여부 확인
    def is_smartstore_product(self, product):
        mall_name = product.get('mallName', '').lower()  # 마켓명
        is_valid = 'smartstore.naver.com' in product.get('mallProductUrl', '').lower()  # 스마트스토어 상품인지 확인
        is_valid = is_valid and all(excluded.lower() not in mall_name for excluded in config['Settings']['excluded_markets'])  # 제외마켓 제외
        logger.debug(f"Product ID {product.get('id', '')} is a smartstore product: {is_valid}")  # 로그
        return is_valid

    # 마켓 정보 추
    def get_market_info(self, product):
        market_name = product.get('mallName', '')  # 마켓
        market_url = product.get('mallPcUrl', '')  # 마켓링크
        return market_name, market_url

    # JSON 다운드 및 저장
    def download_json(self, keyword, sort_type):
        encoded_keyword = keyword.replace(" ", "%20")
        url = f"https://search.shopping.naver.com/search/all?adQuery={encoded_keyword}&origQuery={encoded_keyword}&pagingIndex=1&pagingSize=80&productSet=overseas&query={encoded_keyword}&sort={sort_type}"
        try:
            self.driver.get(url)
            logger.debug(f"URL 접근: {url}")
            time.sleep(random.uniform(2, 5))  # 페이지 로드 대기 시간 추가

            # JSON 데이터 로드
            json_element = self.wait.until(EC.presence_of_element_located((By.ID, "__NEXT_DATA__")))
            json_data = json.loads(json_element.get_attribute('innerHTML'))
            logger.debug("JSON 데이터 로드 성공")

            # JSON 파일 저장
            filename = os.path.join(config['Paths']['download_folder'], f"navershopping_{keyword}_{sort_type}.json")
            with open(filename, 'w', encoding='utf-8') as json_file:
                json.dump(json_data, json_file, ensure_ascii=False, indent=4)
            logger.info(f"JSON 파일 저장 완료: {filename}")
            return filename
        except Exception as e:
            logger.error(f"JSON 데터를 로드하는 중 오류 발생: {str(e)}")
            raise

    # 상품 처리 메서드
    def process_products(self, keyword, products, sort_type_korean):
        processed_products = []
        for item in products:
            product = item.get('item', {})
            if self.is_smartstore_product(product):
                market_name, market_url = self.get_market_info(product)
                is_bundle = '묶음상품' if product.get('mallCount', 0) > 0 else '단품'
                related_keywords = product.get('relatedKeywords', [])
                related_keywords_str = ', '.join(related_keywords) if related_keywords else '없음'
                competitor_count = len(product.get('mallList', []))
                rating_score = product.get('scoreInfo', 0)
                search_trends = product.get('searchTrends', {}).get('trend', '정보 없음')
                rank = product.get('rank', None)

                # 문자열을 정수로 변환
                price = self.safe_int_convert(product.get('price', '0'))
                review_count = self.safe_int_convert(product.get('reviewCount', '0'))
                purchase_count = self.safe_int_convert(product.get('purchaseCnt', '0'))
                recent_purchases = self.safe_int_convert(product.get('recentSales', '0'))

                row = {
                    "id": product.get('id', ''),
                    "keyword": keyword,
                    "market_name": market_name,
                    "product_title": product.get('productTitle', '상품명 없음'),
                    "price": price,
                    "product_url": product.get('mallProductUrl', ''),
                    "market_url": product.get('mallPcUrl', ''),
                    "manu_tag": product.get('manuTag', ''),
                    "image_url": product.get('imageUrl', ''),
                    "category": f"{product.get('category1Name', '')}>{product.get('category2Name', '')}>{product.get('category3Name', '')}",
                    "parentId": product.get('parentId', ''),
                    "category1Id": product.get('category1Id', ''),
                    "category2Id": product.get('category2Id', ''),
                    "category3Id": product.get('category3Id', ''),
                    "category4Id": product.get('category4Id', ''),
                    "brand": product.get('brand', ''),
                    "delivery_price": product.get('dlvryPrice', ''),
                    "is_naver_pay": product.get('isNaverPay', ''),
                    "review_count": review_count,
                    "purchase_count": purchase_count,
                    "recent_purchases": recent_purchases,
                    "overseaTp": product.get('overseaTp', '1'),
                    "delivery_country": product.get('deliveryCountry', ''),
                    "product_type": product.get('type', ''),
                    "sort_type": sort_type_korean,
                    "is_bundle": is_bundle,
                    "related_keywords": related_keywords_str,
                    "search_trends": search_trends,
                    "competitor_count": competitor_count,
                    "rating_score": rating_score,
                    "rank": rank,
                    "mallInfoCache": product.get('mallInfoCache', {})
                }
                processed_products.append(row)
                logger.info(f"Added product data for keyword '{keyword}' with sort type '{sort_type_korean}'")
        return processed_products

    def safe_int_convert(self, value):
        try:
            return int(str(value).replace(',', ''))
        except (ValueError, TypeError):
            return 0

    # JSON 일 저장 메서드
    def save_product(self, data):
        try:
            file_path = os.path.join(self.user_folder, 'collected_products.json')
            existing_data = []
            if os.path.exists(file_path):
                with open(file_path, 'r', encoding='utf-8') as f:
                    existing_data = json.load(f)
            
            # 중복 체크 및 새 데이터 추가
            for item in data:
                if not any(existing_item.get('id') == item.get('id') for existing_item in existing_data):
                    existing_data.append(item)
            
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(existing_data, f, ensure_ascii=False, indent=4)
            logger.debug(f"데이터 저장 완료: {file_path}")
        except Exception as e:
            logger.error(f"데이터 저장 오류: {str(e)}")

# 스크래핑 메인 함수
    def scrape_market(self):
        try:
            self.setup_driver()
            naver_main_url = "https://www.naver.com/"
            self.driver.get(naver_main_url)
            time.sleep(3)  # 초기 페이지 로딩 대기

            all_products = []
            for i in range(1, self.config.get('market', 0) + 1):
                try:
                    option = self.config.get('option', '전체구매수')
                    logger.info(f"옵션 '{option}'으로 마켓 {i}번 처리 중...")
                    market_info = self.config.get('markets', [])[i-1] if i-1 < len(self.config.get('markets', [])) else None
                    if not market_info:
                        logger.warning(f"마켓 정보가 부족합니다: 마켓 {i}")
                        continue
                    market_name = market_info.get('name', f"마{i}")
                    market_url = market_info.get('url', '')
                    if not market_url:
                        logger.warning(f"마켓 URL이 없습니다: {market_name}")
                        continue

                    logger.info(f"{market_name} 마켓 스크래핑 시작: {market_url}")
                    random_delay = random.uniform(3, 6)
                    logger.debug(f"{random_delay:.2f}초 대기")
                    time.sleep(random_delay)

                    scraped_products = self.scrape_single_market(market_url, option)
                    for product in scraped_products:
                        product['collection_method'] = 'market_scraping'
                    all_products.extend(scraped_products)
                except Exception as e:
                    logger.error(f"{market_name} 마켓 스크래핑 중 오류 발생: {str(e)}")
                    continue

            # 상위 10개 상품만 선택
            top_10_products = all_products[:10]
            logger.info(f"상위 10개 상품 추출 완료: {len(top_10_products)}개")

            # 로컬에 저
            saved_file_path = self.store_locally(request.uid, top_10_products)

            return top_10_products, saved_file_path  # 결과 반환
        except Exception as e:
            logger.error(f"검색 중 오류 발생: {str(e)}")
            raise
        finally:
            if self.driver:
                self.driver.quit()
                logger.debug("WebDriver 종료")


    def scrape_market_specific_urls(self, urls, option, keyword=''):
        try:
            self.setup_driver()
            all_results = []
            for url in urls[:3]:  # 최대 3개의 마켓으로 제한
                try:
                    scraped_products = self.scrape_single_market(url, option)
                    all_results.extend(scraped_products)
                except Exception as e:
                    logger.error(f"URL '{url}' 핑 중 류 : {str(e)}")
                    continue
            return all_results
        except Exception as e:
            logger.error(f"마켓 스크래핑 중 오류 발생: {str(e)}")
            raise
        finally:
            if self.driver:
                self.driver.quit()
                logger.debug("WebDriver 종료")



    # 로컬에 저장
    def store_locally(self, uid, products):
        try:
            user_folder = os.path.join(config['Paths']['download_folder'], uid)
            os.makedirs(user_folder, exist_ok=True)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"collected_products_{timestamp}.json"
            file_path = os.path.join(user_folder, filename)

            save_json_data(file_path, products)

            logger.info(f"{len(products)}개의 상품 정보가 {file_path}에 저장되었습니다.")
            return file_path
        except Exception as e:
            logger.error(f"로컬 저장 중 오류 발생: {str(e)}")
            raise

    def collect_selected_products(self, uid, selected_product_ids):
        try:
            user_folder = os.path.join(config['Paths']['download_folder'], uid)  # 사용 폴더 경로 성
            os.makedirs(user_folder, exist_ok=True)  # 사용자 폴더 생성
            selected_products = []  # 선택된 상품 초기화
            market_info = {}  # 켓정보 초화

            # 가장 최근의 collected_products 파일을 찾습니다.
            collected_files = sorted(
                [f for f in os.listdir(user_folder) if f.startswith("collected_products_")],
                reverse=True
            )
            if not collected_files:
                logger.warning(f"사용자 {uid}의 수집 상품 파일 없니다.")
                return {"error": "수집된 상품 정보가 없습니다."}, 404
            latest_file = os.path.join(user_folder, collected_files[0])
            with open(latest_file, 'r', encoding='utf-8') as f:
                products = json.load(f)
            for product in products:
                if product.get('id') in selected_product_ids:
                    selected_products.append(product)
                    # 마켓 정보 추출
                    mall_name = product.get('mallName')  # 마켓명
                    mall_url = product.get('mallPcUrl')  # 마켓링크
                    mall_grade = product.get('mallInfoCache', {}).get('mallGrade', '')  # 마켓등급
                    is_brand_store = product.get('mallInfoCache', {}).get('isBrandStore', False)  # 브랜드스토어여부
                    if mall_name and mall_url:
                        market_info[mall_name] = {
                            'mallName': mall_name,  # 마켓명
                            'mallPcUrl': mall_url,  # 마켓링크
                            'mallGrade': mall_grade,  # 마켓등급
                            'isBrandStore': is_brand_store  # 브드스토어여부
                        }
            if not selected_products:
                logger.warning(f"선택된 상품이 존재하지 않습니다: {selected_product_ids}")
                return {"error": "선택된 상품이 존재하지 않습니다."}, 404
            # 마켓 DB에 저장
            result = self.update_market_db(uid, market_info)

            # 선택된 상품을 로컬에 저장 (DB 구조에 맞게)
            self.save_selected_products(uid, selected_products)

            return {
                "message": f"{len(selected_products)}개의 상품이 수집되었습니다.",
                "market_db_file": result['market_db_file']
            }, 200
        except Exception as e:
            logger.error(f"상품 수집 중 오류 발생: {str(e)}")
            return {"error": "상품 수집 중 오류가 발생했습니다."}, 500

    # 마켓 DB 업데이트
    def update_market_db(self, uid, market_info):
        try:
            user_folder = os.path.join(config['Paths']['download_folder'], uid)  # 사용자 폴더 경로 생성
            os.makedirs(user_folder, exist_ok=True)  # 사용자 폴더 생성
            market_db_file = os.path.join(user_folder, 'market_db.json')  # 마켓 DB 파일 경로 생성

            # 기존 데이터 로드 또는  딕셔너리 생성
            if os.path.exists(market_db_file):  # 마켓 DB 파일이 있으면 로드
                with open(market_db_file, 'r', encoding='utf-8') as f:
                    markets = json.load(f)  # 마켓 DB 로드
            else:  # 마켓 DB 파일이 없으면 새로 생성
                markets = {}
            # 새 데이터 추가 또는 업데이트
            for key, value in market_info.items():  # 마켓정보 추가 또는 업데이트
                markets[key] = value
            # 파일에 저장
            with open(market_db_file, 'w', encoding='utf-8') as f:
                json.dump(markets, f, ensure_ascii=False, indent=4)  # 마켓 DB 저장

            logger.info(f"마켓 DB 업데이트 완료: {market_db_file}")
            return {"market_db_file": market_db_file}
        except Exception as e:
            logger.error(f"마켓 DB 업데이트 중 류 발생: {str(e)}")
            raise

    # 선택된 상을 로컬에 저장
    def save_selected_products(self, uid, selected_products):
        try:
            user_folder = os.path.join(config['Paths']['download_folder'], uid)  # 사용자 폴더 경로 생성
            os.makedirs(user_folder, exist_ok=True)  # 사용자 폴더 생성
            products_db_file = os.path.join(user_folder, 'products_db.json')  # 상품 DB 파일 경로 생성

            # 기존 데이터 로드 또는 새 리스트 생성
            if os.path.exists(products_db_file):
                with open(products_db_file, 'r', encoding='utf-8') as f:
                    products = json.load(f)
            else:
                products = []
            # 새 데이터 추가
            products.extend(selected_products)
            # 파일에 저장
            with open(products_db_file, 'w', encoding='utf-8') as f:
                json.dump(products, f, ensure_ascii=False, indent=4)

            logger.info(f"상품 DB 업데이트 완료: {products_db_file}")
        except Exception as e:
            logger.error(f"상품 DB 업데이트 중 오류 발생: {str(e)}")
            raise

    def scrape_single_market(self, url, option):
            results = []
            try:
                if option == '전체구매건수':
                    url += "/category/ALL?st=TOTALSALE&dt=LIST&page=1&size=80"
                elif option == '3일이내구매건수':
                    url += "/category/ALL?st=POPULAR&dt=LIST&page=1&size=80"
                
                self.driver.get(url)
                logger.debug(f"URL 접근: {url}")
                time.sleep(random.uniform(2, 5))  # 페이지 로드 대기 시간 추가

                # 페이지 로딩 대기
                self.wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "li._3S7Ho5J2Ql")))
                html = self.driver.page_source
                soup = BeautifulSoup(html, 'lxml')
                items = soup.select('li._3S7Ho5J2Ql')

                for index, item in enumerate(items):
                    try:
                        # '전체구매건수' 옵션의 경우 특정 조건에 따라 루프를 종료할 수 있음
                        if option == '전체구매건수':
                            best = item.select_one('div._3lguicci3E > div:nth-child(2) > em')
                            if best:
                                last_best_value = best.text.strip()
                                last_item_with_best = index
                                break  # 필요 시 종료 조건 추가

                        # '3일이내구매건수' 옵션의 경우 'BEST'가 포함된 항목만 처리
                        if option == '3일이내구매건수':
                            elements = item.select_one('li > div.gFNkPpljgY > div._36AX8ncrlx > span > span')
                            if not elements or 'BEST' not in elements.text:
                                continue

                        title = item.select_one('div._1vVKEk_wsi > strong').text.strip()
                        if self.should_skip_title(title):
                            continue

                        price_text = item.select_one('strong._22XUYkkUGJ > span._3_9J443eIx').text
                        price = int(re.sub(r'[^\d]', '', price_text))

                        if option == '전체구매건수':
                            if not (self.config.get('min_price', 0) <= price <= self.config.get('max_price', 1000000000)):
                                continue
                        elif option == '3일이내구매건수':
                            if price < self.config.get('min_price', 0):
                                continue

                        href = item.select_one('#CategoryProducts > ul > li > a')['href']
                        product_url = 'https://smartstore.naver.com' + href

                        elements = item.select_one('li > div.gFNkPpljgY > div._36AX8ncrlx > span > span')
                        elements_text = '(3일이내)' if elements and 'BEST' in elements.text else ''

                        if option == '전체구매건수':
                            purchases = item.select_one('div._3lguicci3E > div:nth-child(2) > em')
                            purchases = purchases.text.strip() if purchases else "0"
                        elif option == '3일이내구매건수':
                            elements = item.select_one('li > div.gFNkPpljgY > div._36AX8ncrlx > span > span')
                            purchases = elements.text.strip() if elements and 'BEST' in elements.text else "0"

                        img = item.select_one('#CategoryProducts > ul > li > div.gFNkPpljgY > div._2JNWBGd-04._3uKZ70Wwcp.EnqpMc_sIs > img')['src']
                        cleaned_url = img.split('?')[0]

                        category = ''
                        category_element = item.select_one('div._3lguicci3E > div:nth-child(1)')
                        if category_element:
                            category = category_element.text.strip()

                        # 3일 이내 구매 건수 추출
                        recent_purchases = int(re.search(r'\d+', purchases).group()) if re.search(r'\d+', purchases) else 0

                        data = {
                            "market_name": url.split("//")[-1].split(".")[0],  # URL에서 마켓명 추출 예시
                            "product_title": title,
                            "price": price,
                            "product_url": product_url,
                            "image_url": cleaned_url,
                            "category": category,
                            "purchases": purchases + '개 이상 ' + elements_text,
                            "recent_purchases": recent_purchases
                        }

                        results.append(data)
                        logger.info(f"Added product: {title} - {price}원, 최근 3일 구매: {recent_purchases}개")
                        time.sleep(random.uniform(1, 2))
                    except Exception as e:
                        logger.error(f"상품 정보 추출  오류 발생: {str(e)}")
                        continue
            except Exception as e:
                logger.error(f"스크래핑 중 오류 발생: {str(e)}")
            return results

    # 마켓 스크래핑 메서드
    def scrape_market_specific_urls(self, urls, option):
        try:
            self.setup_driver()
            all_results = []
            for url in urls[:3]:  # 최대 3개의 마켓으로 제한
                try:
                    scraped_products = self.scrape_single_market(url, option)
                    all_results.extend(scraped_products)
                except Exception as e:
                    logger.error(f"URL '{url}' 스크래핑 중 오류 발생: {str(e)}")
                    continue
            return all_results
        except Exception as e:
            logger.error(f"마켓 스크래핑 중 오류 발생: {str(e)}")
            raise
        finally:
            if self.driver:
                self.driver.quit()
                logger.debug("WebDriver 종료")

    def search_products(self, keyword, uid):
        try:
            self.setup_driver()
            if self.driver is None:
                raise Exception("WebDriver 설정 실패")
            
            all_products = []
            for sort_type in ['review', 'rel']:
                try:
                    json_file = self.download_json(keyword, sort_type)
                    with open(json_file, 'r', encoding='utf-8') as f:
                        json_data = json.load(f)
                
                    products = json_data.get('props', {}).get('pageProps', {}).get('initialState', {}).get('products', {}).get('list', [])
                    sort_type_korean = '리뷰많은순' if sort_type == 'review' else '네이버랭킹순'
                    
                    processed_products = self.process_products(keyword, products, sort_type_korean)
                    all_products.extend(processed_products)
                    
                except Exception as e:
                    logger.error(f"{sort_type} 정렬로 검색 중 오류 발생: {str(e)}")
                    continue

            # 최종적으로 모든 정렬 결과에서 상위 10개 선택
            top_10_products = all_products[:10]
            logger.info(f"최종 상위 10개 상품 추출 완료: {len(top_10_products)}개")

            # 로컬에 저장
            saved_file_path = self.store_locally(uid, top_10_products)

            # MongoDB에 저장
            db.users.update_one(
                {"_id": uid},
                {"$push": {"search_results": {
                    "keyword": keyword,
                    "products": top_10_products,
                    "timestamp": datetime.now()
                }}}
            )

            return top_10_products
        except Exception as e:
            logger.error(f"검색 중 오류 발생: {str(e)}")
            raise
        finally:
            if self.driver:
                self.driver.quit()
                logger.debug("WebDriver 종료")

@app.route('/signup', methods=['POST'])
def signup():
    try:
        data = request.get_json()
        email = data.get('email')
        password = data.get('password')
        name = data.get('name')

        if not email or not password or not name:
            return jsonify({"error": "이메일, 비밀번호, 이름은 필수입니다."}), 400

        if db.users.find_one({"user_info.email": email}):
            return jsonify({"error": "이미 존재하는 이메일입니다."}), 400

        hashed_password = generate_password_hash(password)

        new_user = {
            "_id": str(uuid.uuid4()),
            "user_info": {
                "email": email,
                "password": hashed_password,
                "name": name,
                "membershipLevel": "Premium",
                "remainingCredits": 5
            },
            "config": {
                "market": 0,
                "min_price": 0,
                "max_price": 1000000000,
                "option": "전체구매건수",
                "skip_words": [],
                "markets": []
            },
            "collected_products": [],
            "market_db": {},
            "search_results": []
        }
        
        db.users.insert_one(new_user)

        user_folder = os.path.join(config['Paths']['download_folder'], new_user['_id'])
        os.makedirs(user_folder, exist_ok=True)

        return jsonify({
            "message": "회원가입 성공",
            "uid": new_user['_id']
        }), 201

    except Exception as e:
        logger.error(f"회원가입 중 오류 발생: {str(e)}")
        return jsonify({"error": "회원가입 중 오류가 발생했습니다."}), 500

@app.route('/login', methods=['POST'])
def login():
    try:
        data = request.get_json()
        email = data.get('email')
        password = data.get('password')

        logger.info(f"로그인 시도 - 이메일: {email}")

        user = db.users.find_one({"user_info.email": email})
        if user and check_password_hash(user['user_info']['password'], password):
            logger.info(f"로그인 성공 - 이메일: {email}")
            return jsonify({
                "message": "로그인 성공",
                "uid": str(user['_id']),
                "user": {
                    "name": user['user_info']['name'],
                    "email": user['user_info']['email'],
                    "membershipLevel": user['user_info']['membershipLevel'],
                    "remainingCredits": user['user_info']['remainingCredits']
                }
            }), 200
        else:
            logger.warning(f"로그인 실패 - 이메일: {email}")
            return jsonify({"error": "이메일 또는 비밀번호가 잘못되었습니다."}), 401

    except Exception as e:
        logger.error(f"로그인 중 오류 발생: {str(e)}")
        return jsonify({"error": "로그인 중 오류가 발생했습니다."}), 500

@app.route('/generate_seo', methods=['POST'])
def generate_seo():
    try:
        data = request.get_json(force=True)
        uid = data.get('uid')
        product_id = data.get('product_id')

        user = db.users.find_one({"_id": uid})
        if not user:
            return jsonify({"error": "사용자를 찾을 수 없습니다."}), 404

        product = next((p for p in user['collected_products'] if p['id'] == product_id), None)
        if not product:
            return jsonify({"error": "상품을 찾을 수 없습니다."}), 404

        # 상품명에서 앞 2단어 추출
        words = product['product_title'].split()[:2]
        search_query = ' '.join(words)

        # 네이버 쇼핑 검색
        scraper = NaverShoppingScraper(config['Paths']['download_folder'])
        scraper.setup_driver()
        
        try:
            url = f"https://search.shopping.naver.com/search/all?query={quote(search_query)}&cat_id=&frm=NVSHATC"
            scraper.driver.get(url)
            time.sleep(random.uniform(2, 5))

            # 상품명 추출
            product_titles = scraper.driver.find_elements(By.CSS_SELECTOR, 'div.product_title__Mmw2K')
            titles = [title.text.strip() for title in product_titles[:20]]  # 상위 20개 상품명만 사용

            # 가장 많이 반복되는 단어 추출
            words = ' '.join(titles).split()
            word_counts = Counter(words)
            common_words = [word for word, count in word_counts.most_common(5) if len(word) > 1]

            # SEO 타이틀 생성
            seo_title = ' '.join([search_query] + common_words)
            seo_title = seo_title[:50]  # 50자로 제한

            # MongoDB에 저장
            db.users.update_one(
                {"_id": uid, "collected_products.id": product_id},
                {"$set": {
                    "collected_products.$.seo_title": seo_title,
                }}
            )

            return jsonify({
                "id": product_id,
                "seo_title": seo_title,
            }), 200

        finally:
            scraper.driver.quit()

    except Exception as e:
        logger.error(f"SEO 최적화 중 오류 발생: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({"error": f"SEO 최적화 중 오류가 발생했습니다: {str(e)}"}), 500

import requests

@app.route('/taobao_match', methods=['POST'])
def taobao_match():
    try:
        data = request.get_json(force=True)
        image_url = data.get('image_url')
        
        if not image_url:
            return jsonify({"error": "이미지 URL이 제공되지 않았습니다."}), 400

        url = "https://open-taobao-api.p.rapidapi.com/taobao/traffic/item/imgsearch"
        querystring = {"pic_url": image_url, "language": "en"}
        headers = {
            "x-rapidapi-key": "5d86270c77mshb44bc07820b6f55p1e2c3ajsnfbfd2a9a7943",
            "x-rapidapi-host": "open-taobao-api.p.rapidapi.com"
        }

        response = requests.get(url, headers=headers, params=querystring)
        taobao_data = response.json()

        if taobao_data['code'] == 200 and taobao_data['data']:
            matched_item = taobao_data['data'][0]  # 첫 번째 매칭 아이템 선택
            return jsonify({
                "itemId": matched_item['itemId'],
                "title": matched_item['multiLanguageInfo']['title'],
                "price": matched_item['price'],
                "mainImageUrl": matched_item['mainImageUrl'],
                "shopName": matched_item['shopName']
            }), 200
        else:
            return jsonify({"error": "매칭된 상품이 없습니다."}), 404

    except Exception as e:
        logger.error(f"타오바오 매칭 중 오류 발생: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({"error": "타오바오 매칭 중 오류가 발생했습니다."}), 500

@app.route('/search', methods=['POST', 'OPTIONS'])
def search_products_route():
    if request.method == 'OPTIONS':
        return jsonify({'status': 'OK'}), 200

    try:
        data = request.get_json(force=True)
        keyword = data.get('keyword')
        uid = data.get('uid')

        logger.info(f"검색 요청 받음 - 키워드: {keyword}, UID: {uid}")

        if not uid:
            return jsonify({"error": "유효한 UID가 필요합니다."}), 400

        user = db.users.find_one({"_id": uid})
        if not user:
            return jsonify({"error": "사용��� 찾을 수 습���."}), 404

        scraper = NaverShoppingScraper(config['Paths']['download_folder'])
        top_10_products = scraper.search_products(keyword, uid)  # uid 전달

        return jsonify({
            "products": top_10_products
        }), 200

    except Exception as e:
        logger.error(f"검색 중 오류 발생: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({"error": "검색 중 오류가 발생했습니다."}), 500


@app.route('/get_collected_products', methods=['GET'])
def get_collected_products():
    try:
        uid = request.args.get('uid')
        if not uid:
            return jsonify({"error": "사용자 ID가 필요합니다."}), 400

        user = db.users.find_one({"_id": uid})
        if not user:
            return jsonify({"error": "사용자를 찾을 수 없습니다."}), 404

        collected_products = user.get('collected_products', [])
        
        # seo_title 필드가 없는 경우 빈 문자열로 초기화
        for product in collected_products:
            if 'seo_title' not in product:
                product['seo_title'] = ''

        return jsonify({"products": collected_products}), 200
    except Exception as e:
        logger.error(f"수집된 상품 조회 중 오류 발생: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({"error": f"수집된 상품 조회 중 오류가 발생했습니다: {str(e)}"}), 500


@app.route('/batch_taobao_match', methods=['POST'])
def batch_taobao_match():
    try:
        data = request.get_json(force=True)
        uid = data.get('uid')
        product_ids = data.get('productIds', [])
        
        if not uid or not product_ids:
            return jsonify({"error": "사용자 ID와 매칭할 상품 ID가 필요합니다."}), 400

        user = db.users.find_one({"_id": uid})
        if not user:
            return jsonify({"error": "사용자를 찾을 수 없습니다."}), 404

        collected_products = user.get('collected_products', [])
        products_to_match = [p for p in collected_products if p['id'] in product_ids]

        matched_products = []
        for product in products_to_match:
            image_url = product.get('image_url')
            if not image_url:
                continue

            url = "https://open-taobao-api.p.rapidapi.com/taobao/traffic/item/imgsearch"
            querystring = {"pic_url": image_url, "language": "en"}
            headers = {
                "x-rapidapi-key": "5d86270c77mshb44bc07820b6f55p1e2c3ajsnfbfd2a9a7943",
                "x-rapidapi-host": "open-taobao-api.p.rapidapi.com"
            }

            response = requests.get(url, headers=headers, params=querystring)
            taobao_data = response.json()

            if taobao_data['code'] == 200 and taobao_data['data']:
                matched_item = taobao_data['data'][0]
                product['taobaoMatch'] = {
                    "itemId": matched_item['itemId'],
                    "title": matched_item['multiLanguageInfo']['title'],
                    "price": matched_item['price'],
                    "mainImageUrl": matched_item['mainImageUrl'],
                    "shopName": matched_item['shopName']
                }
            matched_products.append(product)

        # Update the user's collected_products in the database
        db.users.update_one(
            {"_id": uid},
            {"$set": {"collected_products": matched_products}}
        )

        return jsonify({"matched_products": matched_products}), 200

    except Exception as e:
        logger.error(f"일괄 타오바오 매칭 중 오류 발생: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({"error": "일괄 타오바오 매칭 중 오류가 발생했습니다."}), 500

@app.route('/user-info', methods=['GET'])
def get_user_info():
    uid = request.args.get('uid')
    if not uid:
        return jsonify({"error": "사용자 ID가 필요합니다."}), 400

    user = db.users.find_one({"_id": uid})
    if user:
        return jsonify({
            "user": {
                "name": user['user_info']['name'],
                "email": user['user_info']['email'],
                "membershipLevel": user['user_info']['membershipLevel'],
                "remainingCredits": user['user_info']['remainingCredits']
            }
        }), 200
    else:
        return jsonify({"error": "사용자를 찾을 수 없습니다."}), 404

@app.route('/collect', methods=['POST'])
def collect_products():
    try:
        data = request.get_json(force=True)
        uid = data.get('uid')
        selected_product_ids = data.get('selected_product_ids', [])

        if not uid or not selected_product_ids:
            return jsonify({"error": "사용자 ID와 선택된 상품 ID가 필요합니다."}), 400

        user = db.users.find_one({"_id": uid})
        if not user:
            return jsonify({"error": "사용자를 찾을 수 없습니다."}), 404

        search_results = user.get('search_results', [])
        if not search_results:
            return jsonify({"error": "검색 결과가 없습니다."}), 404

        latest_search_result = search_results[-1]['products']
        collected_products = [
            {**product, 'collection_method': 'keyword_search'} 
            for product in latest_search_result 
            if str(product['id']) in selected_product_ids
        ]

        db.users.update_one(
            {"_id": uid},
            {"$push": {"collected_products": {"$each": collected_products}}}
        )

        return jsonify({"message": f"{len(collected_products)}개의 상품이 성공적으로 수집되었습니다."}), 200

    except Exception as e:
        logger.error(f"상품 수집 중 오류 발생: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({"error": "상품 수집 중 오류가 발생했습니다."}), 500

@app.route('/get_market_db', methods=['GET'])
def get_market_db_route():
    uid = request.uid

    try:
        user_folder = os.path.join(config['Paths']['download_folder'], uid)
        market_db_file = os.path.join(user_folder, 'market_db.json')

        if os.path.exists(market_db_file):
            with open(market_db_file, 'r', encoding='utf-8') as f:
                markets = json.load(f)
            
            # markets가 리스트인 경우 그대로 반환, 딕셔너리인 경우 values()를 사용
            if isinstance(markets, list):
                return jsonify({"markets": markets}), 200
            elif isinstance(markets, dict):
                return jsonify({"markets": list(markets.values())}), 200
            else:
                logger.error(f"Unexpected market data type: {type(markets)}")
                return jsonify({"error": "마켓 데이터 형식이 올바르지 않습니다."}), 500
        else:
            return jsonify({"markets": []}), 200
    except Exception as e:
        logger.error(f"마켓 DB 조회 중 오류 발생: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({"error": "마켓 DB 조회 중 오류가 발생했습니다."}), 500

@app.route('/collect_market', methods=['POST'])
def collect_market_route():
    try:
        data = request.get_json(force=True)
        uid = request.uid
        market_data = data.get('market_data', [])

        if not market_data or not isinstance(market_data, list):
            return jsonify({"error": "유효한 마켓 데이터가 제공되지 않았습니다."}), 400

        user_folder = os.path.join(config['Paths']['download_folder'], uid)
        os.makedirs(user_folder, exist_ok=True)
        market_db_file = os.path.join(user_folder, 'market_db.json')

        # 기존 마켓 DB 로드 또는 새로 생성
        existing_markets = load_json_data(market_db_file)

        if not isinstance(existing_markets, list):
            existing_markets = []

        new_markets_added = 0
        for new_market in market_data:
            if isinstance(new_market, dict) and 'mallName' in new_market:
                if not any(market.get('mallName') == new_market['mallName'] for market in existing_markets):
                    existing_markets.append(new_market)
                    new_markets_added += 1

        # 업데이트된 마켓 DB 저장
        save_json_data(market_db_file, existing_markets)

        return jsonify({"message": f"{new_markets_added}개의 새로운 마켓 정보가 성공적으로 저장되었습니다."}), 200
    except Exception as e:
        logger.error(f"마켓 수집 요청 처리 중 오류 발생: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({"error": f"마켓 수집 요청 처리 중 오류가 발생했습니다: {str(e)}"}), 500
    


@app.route('/post_process', methods=['POST'])
def post_process_route():
    try:
        data = request.get_json(force=True)
        uid = request.uid

        scraper = NaverShoppingScraper(os.path.join(app.config['Paths']['download_folder'], uid))
        scraper.post_process_data(uid)

        return jsonify({"message": "데이터 후처리가 완료되었습니다."}), 200
    except Exception as e:
        logger.error(f"데이터 후처리 요청 중 오류 발생: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({"error": "데이터 후처리 요청 중 오류가 발생했습니다."}), 500

@app.route('/download_heyseller', methods=['GET'])
def download_heyseller():
    try:
        uid = request.args.get('uid')
        if not uid:
            return jsonify({"error": "유효한 UID가 필요합니다."}), 400

        user = db.users.find_one({"_id": uid})
        if not user:
            return jsonify({"error": "사용자를 찾을 수 없습니다."}), 404

        # 템플릿 파일 가져오기
        template = db.templates.find_one({"name": "heyseller_template"})
        if not template or 'file' not in template:
            return jsonify({"error": "헤이셀러 템플릿을 찾을 수 없습니다."}), 404

        # 템플릿 파일 로드
        template_file = io.BytesIO(template['file'])
        workbook = load_workbook(template_file)
        sheet = workbook.active

        collected_products = user.get('collected_products', [])
        
        # 템플릿에 데이터 채우기
        for idx, product in enumerate(collected_products, start=3):  # 3행부터 시작
            sheet.cell(row=idx, column=1, value=product.get('category', ''))  # A열: 카테고리
            
            # B열: SEO 상품명 (50자 이하로 제한)
            seo_title = product.get('seo_title', product['product_title'])
            seo_title = seo_title[:50]  # 50자로 제한
            sheet.cell(row=idx, column=2, value=seo_title)
            
            sheet.cell(row=idx, column=3, value=len(seo_title.encode('utf-8')))  # C열: 글자수(Byte)
            
            # D열: 배대지 비용 (0으로 설정)
            sheet.cell(row=idx, column=4, value=0)
            
            # E열: 타오바오 URL
            taobao_id = product.get('taobaoMatch', {}).get('itemId', '')
            taobao_url = f"https://item.taobao.com/item.htm?id={taobao_id}" if taobao_id else ''
            sheet.cell(row=idx, column=5, value=taobao_url)
            
            # F열은 비워둠
            sheet.cell(row=idx, column=6, value='')

        # 메모리에 엑셀 파일 저장
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name=f"헤이셀러_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logger.error(f"헤이셀러 다운로드 중 오류 발생: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({"error": f"헤이셀러 다운로드 중 오류가 발생했습니다: {str(e)}"}), 500

# Flask 애플리케이션 실행
if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)